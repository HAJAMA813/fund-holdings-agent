from __future__ import annotations

import json
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from fund_holdings_agent.excel_reports import audit_workbook, build_manager_fund_pool_report


def test_fund_pool_report_is_a_standalone_openpyxl_workbook(tmp_path: Path) -> None:
    source = tmp_path / "manager_fund_pool_data.json"
    output = tmp_path / "基金池.xlsx"
    source.write_text(
        json.dumps(
            {
                "summary": {
                    "manager": "测试经理",
                    "manager_id": "123",
                    "company": "测试基金",
                    "report_date": "2026-06-30",
                    "historical_share_count": 1,
                    "active_share_count": 1,
                    "selected_share_count": 1,
                    "product_count": 1,
                    "verified_count": 1,
                    "error_count": 0,
                    "warning_count": 0,
                },
                "selected_funds": [
                    {
                        "fund_code": "000001",
                        "fund_name": "测试基金A",
                        "product_base_name": "测试基金",
                        "product_group": "测试基金",
                        "fund_type": "混合型",
                        "tenure_start": "2024-01-01",
                        "tenure_end": "",
                        "active_on_report_date": True,
                        "inception_date": "2023-01-01",
                        "verified_manager": "测试经理",
                        "manager_verification": "通过",
                        "selected": True,
                        "selection_reason": "报告期在任",
                    }
                ],
                "all_tenures": [],
                "issues": [],
                "sources": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    build_manager_fund_pool_report(source, output)

    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["运行摘要", "报告期基金池", "全部任职历史", "核验与异常", "来源与口径"]
    assert workbook["报告期基金池"]["A2"].value == "000001"
    assert workbook["报告期基金池"]["A2"].number_format == "@"
    assert workbook["报告期基金池"]["F2"].number_format == "yyyy-mm-dd"
    fund_pool_sheet = workbook["报告期基金池"]
    assert len(fund_pool_sheet.tables) == 1
    assert fund_pool_sheet.auto_filter.ref is None

    with ZipFile(output) as archive:
        sheet_xml = archive.read("xl/worksheets/sheet2.xml")
        table_xml = archive.read("xl/tables/table1.xml")
    assert b"<autoFilter" not in sheet_xml
    assert b"<autoFilter" in table_xml
    assert audit_workbook(output)["valid"] is True
