from __future__ import annotations

import datetime as dt
import json
from pathlib import Path

from openpyxl import load_workbook

from fund_holdings_agent.excel_reports import audit_workbook, build_three_quarter_brief_report
from fund_holdings_agent.pdf_reports import build_three_quarter_pdf_report
from fund_holdings_agent.three_quarter import build_three_quarter_dataset, discover_quarter_inputs, save_three_quarter_dataset, three_quarter_dates


def _write_quarter(root: Path, manager: str, date: dt.date, holdings: list[dict], products: list[dict]) -> None:
    quarter = f"{date.year}Q{date.month // 3}"
    directory = root / f"{manager}_{quarter}"
    directory.mkdir(parents=True)
    (directory / "manager_fund_pool_data.json").write_text(
        json.dumps(
            {
                "summary": {"manager": manager, "company": "测试基金管理有限公司", "report_date": date.isoformat(), "selected_share_count": len(products)},
                "selected_funds": products,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (directory / "industry_analysis_data.json").write_text(
        json.dumps(
            {
                "summary": {"manager": manager, "company": "测试基金管理有限公司", "report_date": date.isoformat(), "formal_holding_rows": len(holdings)},
                "formal_holdings_industry": holdings,
                "issues": [],
                "industry_issues": [],
                "industry_quality": {"holding_coverage": 1.0, "snapshot_date": "2026-08-16"},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def test_three_quarter_dates_are_contiguous() -> None:
    assert three_quarter_dates(dt.date(2026, 6, 30)) == [
        dt.date(2025, 12, 31),
        dt.date(2026, 3, 31),
        dt.date(2026, 6, 30),
    ]


def test_builds_three_quarter_dataset_and_brief_workbook(tmp_path: Path) -> None:
    manager = "测试经理"
    products = [
        {"fund_code": "000001", "fund_name": "测试成长A", "product_base_name": "测试成长", "selected": True},
        {"fund_code": "000002", "fund_name": "测试成长C", "product_base_name": "测试成长", "selected": True},
    ]
    dates = three_quarter_dates(dt.date(2026, 6, 30))
    for index, date in enumerate(dates):
        _write_quarter(
            tmp_path,
            manager,
            date,
            [
                {"fund_code": "000001", "fund_name": "测试成长A", "rank": 1, "stock_code": f"60000{index}.SH", "stock_name": f"股票{index}", "market": "A股", "sw_level1": "电子", "shares_10k": 10 + index, "market_value_10k": 100 + index, "nav_ratio": 0.05 + index / 100},
                {"fund_code": "000001", "fund_name": "测试成长A", "rank": 2, "stock_code": "00700.HK", "stock_name": "港股样本", "market": "港股", "sw_level1": "申万不适用"},
            ],
            products,
        )

    inputs = discover_quarter_inputs(tmp_path, manager, dates[-1])
    data = build_three_quarter_dataset(inputs)
    assert data["summary"]["quarters"] == ["2025Q4", "2026Q1", "2026Q2"]
    assert data["summary"]["product_count"] == 1
    assert data["summary"]["a_share_holding_rows"] == 3
    assert data["summary"]["non_a_holding_rows_excluded"] == 3
    assert len(data["rows"]) == 10
    assert data["rows"][0]["quarters"]["2026Q2"]["stock_name"] == "股票2"
    assert data["rows"][0]["quarters"]["2026Q2"]["nav_ratio"] == 0.07
    assert data["rows"][1]["quarters"]["2026Q2"] == {}
    assert data["analytics"]["stock_summary_by_quarter"]["2026Q2"][0]["market_value_10k_sum"] == 102.0
    assert data["analytics"]["industry_summary_by_quarter"]["2026Q2"][0]["nav_ratio_sum"] == 0.07
    assert data["analytics"]["stock_changes"]["new_count"] == 1
    assert data["analytics"]["stock_changes"]["exited_count"] == 1

    json_path = save_three_quarter_dataset(data, tmp_path / "brief.json")
    output = build_three_quarter_brief_report(json_path, tmp_path / "brief.xlsx")
    workbook = load_workbook(output, data_only=False)
    assert workbook.sheetnames == ["01_三季持仓", "99_说明异常"]
    assert workbook["01_三季持仓"]["D5"].value == "2025Q4持仓"
    assert workbook["01_三季持仓"]["H5"].value == "2026Q2持仓"
    assert workbook["01_三季持仓"]["H6"].value == "股票2"
    assert workbook["01_三季持仓"]["H7"].value is None
    assert workbook["99_说明异常"]["B10"].value == "未调用"
    assert audit_workbook(output, expected_sheets=["01_三季持仓", "99_说明异常"])["valid"] is True

    pdf_output = build_three_quarter_pdf_report(json_path, tmp_path / "brief.pdf")
    assert pdf_output.read_bytes().startswith(b"%PDF")
    assert pdf_output.stat().st_size > 5_000
