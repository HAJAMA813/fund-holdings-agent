from __future__ import annotations

import csv
import datetime as dt
import json
import subprocess
import termios
from pathlib import Path

from openpyxl import load_workbook

from fund_holdings_agent import mac_cli
from fund_holdings_agent.mac_cli import (
    MacAgentPaths,
    build_portfolio_command,
    default_paths,
    initialize,
    list_manager_runs,
    list_results,
    list_status,
    run_three_quarter_brief,
    select_roster,
)


def _write_business_inputs(root: Path) -> None:
    data = root / "data"
    data.mkdir(parents=True)
    with (data / "managers_portfolio.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["company", "manager", "manager_id", "active"])
        writer.writeheader()
        writer.writerow({"company": "甲基金管理有限公司", "manager": "甲经理", "manager_id": "100", "active": "yes"})
        writer.writerow({"company": "乙基金管理有限公司", "manager": "乙经理", "manager_id": "200", "active": "yes"})
    (data / "personnel_internal_20260616.csv").write_text("placeholder\n", encoding="utf-8")
    (data / "resource_candidate_confirmations.csv").write_text("placeholder\n", encoding="utf-8")


def test_default_paths_support_isolated_mac_user_directories(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("FUND_AGENT_APP_HOME", str(tmp_path / "app"))
    monkeypatch.setenv("FUND_AGENT_REPORT_ROOT", str(tmp_path / "reports"))
    paths = default_paths()
    assert paths.app_home == (tmp_path / "app").resolve()
    assert paths.report_root == (tmp_path / "reports").resolve()


def test_initialize_and_select_roster_without_touching_real_user_home(tmp_path: Path) -> None:
    root = tmp_path / "project"
    _write_business_inputs(root)
    paths = MacAgentPaths(tmp_path / "app", tmp_path / "reports")
    config = initialize(root, paths)
    selected_path = paths.state_root / "selected.csv"
    selected = select_roster(Path(config["roster"]), selected_path, managers=["乙经理"])
    assert [row.manager for row in selected] == ["乙经理"]
    assert paths.config_path.exists()
    assert "乙经理" in selected_path.read_text(encoding="utf-8-sig")


def test_build_command_uses_stable_user_directories_and_data_only_mode(tmp_path: Path) -> None:
    config = {
        "portfolio_root": str(tmp_path / "portfolio"),
        "report_root": str(tmp_path / "reports"),
        "personnel": str(tmp_path / "personnel.csv"),
        "candidate_confirmations": str(tmp_path / "confirmations.csv"),
    }
    command = build_portfolio_command(
        config,
        tmp_path / "selected.csv",
        report_date="2026-06-30",
        retry_errors=True,
        refresh=False,
        data_only=True,
    )
    assert "2026-06-30" in command
    assert "--retry-errors" in command
    assert "--skip-reports" in command
    assert "--skip-company-report" in command
    assert "--node" not in command
    assert "--friendly-output" in command


def test_status_and_results_read_existing_user_outputs(tmp_path: Path) -> None:
    portfolio = tmp_path / "portfolio"
    reports = tmp_path / "reports"
    portfolio.mkdir()
    reports.mkdir()
    (portfolio / "portfolio_summary_2026Q2.json").write_text(
        json.dumps({"report_date": "2026-06-30", "overall_status": "completed", "manager_results": [{"manager": "甲"}], "notification_summary": "ok"}),
        encoding="utf-8",
    )
    workbook = reports / "甲基金_2026Q2.xlsx"
    workbook.write_bytes(b"xlsx")
    pdf = reports / "甲基金_2026Q2.pdf"
    pdf.write_bytes(b"pdf")
    config = {"portfolio_root": str(portfolio), "report_root": str(reports)}
    assert list_status(config)[0]["quarter"] == "2026Q2"
    assert set(list_results(config)) == {workbook, pdf}


def test_manager_run_status_reads_in_progress_manifests(tmp_path: Path) -> None:
    portfolio = tmp_path / "portfolio"
    manager_dir = portfolio / "甲基金" / "甲经理_2026Q2"
    manager_dir.mkdir(parents=True)
    (manager_dir / "batch_manifest.json").write_text(
        json.dumps({
            "manager": "甲经理",
            "report_date": "2026-06-30",
            "overall_status": "running",
            "updated_at": "2026-08-15T22:00:00+08:00",
            "stages": {"fund_pool": {"status": "completed"}, "holdings": {"status": "running"}},
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    rows = list_manager_runs({"portfolio_root": str(portfolio)})
    assert rows[0]["manager"] == "甲经理"
    assert rows[0]["stage"] == "holdings"


def _write_brief_quarter(portfolio_root: Path, report_date: dt.date) -> None:
    quarter = f"{report_date.year}Q{report_date.month // 3}"
    directory = portfolio_root / "甲基金" / f"甲经理_{quarter}"
    directory.mkdir(parents=True, exist_ok=True)
    (directory / "manager_fund_pool_data.json").write_text(
        json.dumps(
            {
                "summary": {"manager": "甲经理", "company": "甲基金管理有限公司", "report_date": report_date.isoformat(), "selected_share_count": 1},
                "selected_funds": [{"fund_code": "000001", "fund_name": "甲成长A", "product_base_name": "甲成长", "selected": True}],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (directory / "industry_analysis_data.json").write_text(
        json.dumps(
            {
                "summary": {"manager": "甲经理", "company": "甲基金管理有限公司", "report_date": report_date.isoformat(), "formal_holding_rows": 1},
                "formal_holdings_industry": [{"fund_code": "000001", "fund_name": "甲成长A", "rank": 1, "stock_code": "600000.SH", "stock_name": quarter, "market": "A股", "sw_level1": "银行", "market_value_10k": 100, "nav_ratio": 0.05}],
                "issues": [],
                "industry_issues": [],
                "industry_quality": {"holding_coverage": 1.0, "snapshot_date": "2026-08-16"},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (directory / "batch_manifest.json").write_text(json.dumps({"overall_status": "completed"}), encoding="utf-8")


def test_three_quarter_brief_reuses_existing_quarters_and_builds_excel(tmp_path: Path, monkeypatch) -> None:
    project = tmp_path / "project"
    _write_business_inputs(project)
    paths = MacAgentPaths(tmp_path / "app", tmp_path / "reports")
    config = initialize(project, paths)
    for date in (dt.date(2025, 12, 31), dt.date(2026, 3, 31), dt.date(2026, 6, 30)):
        _write_brief_quarter(paths.portfolio_root, date)

    monkeypatch.setattr(mac_cli, "doctor", lambda config: {"status": "RUNNABLE", "checks": [], "note": "ok"})
    monkeypatch.setattr(mac_cli.subprocess, "run", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("已有季度不应补跑")))

    code = run_three_quarter_brief(
        config,
        paths,
        end_report_date="2026-06-30",
        companies=[],
        managers=["甲经理"],
        refresh=False,
    )

    assert code == 0
    output = paths.report_root / "甲基金" / "甲基金_甲经理_2025Q4-2026Q2_三季度前十大持仓简报.xlsx"
    assert output.exists()
    assert load_workbook(output).sheetnames == ["01_三季持仓", "99_说明异常"]


def test_three_quarter_brief_can_build_excel_and_pdf(tmp_path: Path, monkeypatch) -> None:
    project = tmp_path / "project"
    _write_business_inputs(project)
    paths = MacAgentPaths(tmp_path / "app", tmp_path / "reports")
    config = initialize(project, paths)
    for date in (dt.date(2025, 12, 31), dt.date(2026, 3, 31), dt.date(2026, 6, 30)):
        _write_brief_quarter(paths.portfolio_root, date)

    monkeypatch.setattr(mac_cli, "doctor", lambda config: {"status": "RUNNABLE", "checks": [], "note": "ok"})
    monkeypatch.setattr(mac_cli.subprocess, "run", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("已有季度不应补跑")))

    code = run_three_quarter_brief(
        config,
        paths,
        end_report_date="2026-06-30",
        companies=[],
        managers=["甲经理"],
        refresh=False,
        output_format="both",
    )

    assert code == 0
    root = paths.report_root / "甲基金"
    assert (root / "甲基金_甲经理_2025Q4-2026Q2_三季度前十大持仓简报.xlsx").exists()
    pdf = root / "甲基金_甲经理_2025Q4-2026Q2_三季度持仓分析报告.pdf"
    assert pdf.read_bytes().startswith(b"%PDF")


def test_three_quarter_brief_only_runs_missing_quarter(tmp_path: Path, monkeypatch) -> None:
    project = tmp_path / "project"
    _write_business_inputs(project)
    paths = MacAgentPaths(tmp_path / "app", tmp_path / "reports")
    config = initialize(project, paths)
    _write_brief_quarter(paths.portfolio_root, dt.date(2025, 12, 31))
    _write_brief_quarter(paths.portfolio_root, dt.date(2026, 3, 31))
    commands: list[list[str]] = []

    def fake_run(command, **kwargs):
        commands.append(command)
        assert command[command.index("--report-date") + 1] == "2026-06-30"
        _write_brief_quarter(paths.portfolio_root, dt.date(2026, 6, 30))
        return subprocess.CompletedProcess(command, 0)

    monkeypatch.setattr(mac_cli, "doctor", lambda config: {"status": "RUNNABLE", "checks": [], "note": "ok"})
    monkeypatch.setattr(mac_cli.subprocess, "run", fake_run)

    code = run_three_quarter_brief(
        config,
        paths,
        end_report_date="2026-06-30",
        companies=[],
        managers=["甲经理"],
        refresh=False,
    )

    assert code == 0
    assert len(commands) == 1


def test_discard_pending_input_ignores_non_tty_termios_error(monkeypatch) -> None:
    monkeypatch.setattr(termios, "tcflush", lambda *args: (_ for _ in ()).throw(termios.error(19, "not supported")))
    mac_cli._discard_pending_input()


def test_scope_manager_names_filters_roster(tmp_path: Path) -> None:
    roster = tmp_path / "roster.csv"
    roster.write_text(
        "company,manager,manager_id,active\n甲基金管理有限公司,甲经理,1,yes\n甲基金管理有限公司,乙经理,2,yes\n乙基金管理有限公司,丙经理,3,yes\n",
        encoding="utf-8",
    )
    config = {"roster": str(roster)}
    assert mac_cli._scope_manager_names(config, [], ["乙经理"]) == ["乙经理"]
    assert set(mac_cli._scope_manager_names(config, ["甲基金管理有限公司"], [])) == {"甲经理", "乙经理"}


def test_llm_interpretation_helpers_skip_without_key(tmp_path: Path, monkeypatch) -> None:
    from fund_holdings_agent import llm_agent

    monkeypatch.setattr(llm_agent, "llm_available", lambda: False)
    config = {
        "roster": str(tmp_path / "roster.csv"),
        "portfolio_root": str(tmp_path / "portfolio"),
        "report_root": str(tmp_path / "reports"),
    }
    assert mac_cli._llm_interpret_quarter(config, "甲经理", "2026-06-30") is False
    assert mac_cli._llm_interpret_status(config) is False
