from __future__ import annotations

import datetime as dt
import json
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


NAVY = "17365D"
TEAL = "1F7A8C"
PALE = "EAF2F8"
GREEN = "E2F0D9"
GRAY = "F3F6F9"
WHITE = "FFFFFF"
TEXT = "172B4D"
LIGHT_BORDER = Side(style="thin", color="D9E2EC")
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
FORMULA_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")

Column = tuple[str, str | Callable[[dict[str, Any], int], Any], str | None, float | None]


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _new_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    return workbook


def _safe_sheet_name(value: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", value)[:31]


def _table_name(workbook: Workbook, sheet_name: str) -> str:
    count = sum(len(sheet.tables) for sheet in workbook.worksheets) + 1
    stem = re.sub(r"[^A-Za-z0-9_]", "", sheet_name) or "Report"
    return f"T{count:02d}_{stem[:20]}"


def _excel_value(value: Any, header: str = "") -> Any:
    if value is None:
        return ""
    if isinstance(value, (bool, int, float, dt.date, dt.datetime)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return ""
        return value
    text = str(value)
    if DATE_RE.match(text) and any(token in header for token in ("日期", "报告期", "报告日", "任职", "成立", "快照")):
        return dt.date.fromisoformat(text)
    return text


def _style_title(ws, title: str, last_col: int) -> None:
    last_col = max(last_col, 8)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30


def _style_header(cells: Iterable[Any]) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=LIGHT_BORDER)


def _format_body(ws, start_row: int, end_row: int, last_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.font = Font(name="Arial", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(bottom=LIGHT_BORDER)
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="EAF7FB")


def _set_widths(ws, headers: Sequence[str], rows: Sequence[Sequence[Any]], explicit: dict[str, float] | None = None) -> None:
    explicit = explicit or {}
    for index, header in enumerate(headers, start=1):
        if header in explicit:
            width = explicit[header]
        else:
            samples = [str(row[index - 1]) for row in rows[:150] if index - 1 < len(row) and row[index - 1] not in (None, "")]
            longest = max([len(header), *(min(len(value), 40) for value in samples)], default=len(header))
            width = min(32, max(10, longest * 1.15 + 2))
        ws.column_dimensions[get_column_letter(index)].width = width


def _number_format(header: str) -> str | None:
    if any(token in header for token in ("比例", "覆盖率", "变化率")):
        return "0.00%"
    if any(token in header for token in ("市值", "万股")):
        return "#,##0.00"
    if any(token in header for token in ("数量", "次数", "记录数", "基金数", "经理数", "排名", "匹配分", "错误", "警告")):
        return "#,##0"
    if any(token in header for token in ("日期", "报告期", "报告日", "任职", "成立", "快照")):
        return "yyyy-mm-dd"
    return None


def add_table_sheet(
    workbook: Workbook,
    name: str,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    widths: dict[str, float] | None = None,
    title_row: bool = True,
) -> Any:
    ws = workbook.create_sheet(_safe_sheet_name(name))
    header_row = 2 if title_row else 1
    if title_row:
        _style_title(ws, title, len(headers))
    for column, header in enumerate(headers, start=1):
        ws.cell(header_row, column, header)
    _style_header(ws[header_row])
    ws.row_dimensions[header_row].height = 30
    normalized: list[list[Any]] = []
    for raw in rows:
        row = [_excel_value(value, headers[index]) for index, value in enumerate(raw)]
        normalized.append(row)
        ws.append(row)
    data_start = header_row + 1
    data_end = header_row + len(normalized)
    if normalized:
        _format_body(ws, data_start, data_end, len(headers))
        ref = f"A{header_row}:{get_column_letter(len(headers))}{data_end}"
        table = Table(displayName=_table_name(workbook, name), ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    else:
        # A worksheet-level AutoFilter may not overlap an Excel Table's own
        # AutoFilter. Excel for Mac repairs such files by deleting the table.
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
    ws.freeze_panes = f"A{data_start}"
    ws.sheet_view.showGridLines = False
    _set_widths(ws, headers, normalized, widths)
    for column, header in enumerate(headers, start=1):
        fmt = _number_format(header)
        if fmt and data_end >= data_start:
            for cell in ws.iter_cols(min_col=column, max_col=column, min_row=data_start, max_row=data_end):
                for value_cell in cell:
                    value_cell.number_format = fmt
    return ws


def _rows(data: Sequence[dict[str, Any]], columns: Sequence[Column]) -> list[list[Any]]:
    result: list[list[Any]] = []
    for index, row in enumerate(data, start=2):
        values = []
        for _, key, _, _ in columns:
            values.append(key(row, index) if callable(key) else row.get(key, ""))
        result.append(values)
    return result


def _headers(columns: Sequence[Column]) -> list[str]:
    return [row[0] for row in columns]


def _widths(columns: Sequence[Column]) -> dict[str, float]:
    return {label: width for label, _, _, width in columns if width}


def _apply_column_formats(ws, columns: Sequence[Column], start_row: int, end_row: int) -> None:
    for column_index, (_, _, fmt, _) in enumerate(columns, start=1):
        if fmt:
            for row in range(start_row, end_row + 1):
                ws.cell(row, column_index).number_format = fmt


def add_summary_sheet(
    workbook: Workbook,
    title: str,
    left: Sequence[tuple[str, Any]],
    right: Sequence[tuple[str, Any]],
    *,
    note: str,
    right_title: str = "质量指标",
) -> Any:
    ws = workbook.create_sheet("运行摘要")
    _style_title(ws, title, 8)
    ws.cell(3, 1, left[0][0] if left else "指标")
    ws.cell(3, 2, left[0][1] if left else "")
    for offset, (label, value) in enumerate(left[1:], start=4):
        ws.cell(offset, 1, label)
        ws.cell(offset, 2, _excel_value(value, label))
    ws.cell(3, 4, right_title)
    ws.cell(3, 5, "数量/结论")
    for offset, (label, value) in enumerate(right, start=4):
        ws.cell(offset, 4, label)
        ws.cell(offset, 5, _excel_value(value, label))
    _style_header(ws[3][3:5])
    max_row = max(3 + len(left) - 1, 3 + len(right))
    for row in range(3, max_row + 1):
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=PALE)
        ws.cell(row, 1).font = Font(name="Arial", bold=True, color=TEXT)
        ws.cell(row, 4).font = Font(name="Arial", color=TEXT)
        for col in (1, 2, 4, 5):
            ws.cell(row, col).border = Border(bottom=LIGHT_BORDER)
    note_row = max_row + 3
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 1, end_column=8)
    note_cell = ws.cell(note_row, 1, note)
    note_cell.fill = PatternFill("solid", fgColor=GRAY)
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    note_cell.font = Font(name="Arial", size=10, color=TEXT)
    ws.row_dimensions[note_row].height = 34
    for col, width in {1: 25, 2: 24, 3: 4, 4: 24, 5: 24, 6: 4, 7: 4, 8: 4}.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False
    return ws


def _save(workbook: Workbook, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_suffix(output_path.suffix + ".tmp")
    workbook.save(temporary)
    temporary.replace(output_path)
    return output_path


def _move_summary_first(workbook: Workbook) -> None:
    summary = workbook["运行摘要"]
    workbook._sheets.remove(summary)
    workbook._sheets.insert(0, summary)


def build_manager_fund_pool_report(input_path: Path, output_path: Path) -> Path:
    data = read_json(input_path)
    summary = data["summary"]
    workbook = _new_workbook()
    columns: list[Column] = [
        ("基金代码", "fund_code", "@", 12), ("基金名称", "fund_name", None, 26),
        ("基础产品名称", "product_base_name", None, 24), ("产品组", "product_group", None, 14),
        ("基金类型", "fund_type", None, 18), ("任职开始", "tenure_start", "yyyy-mm-dd", 13),
        ("任职结束", "tenure_end", "yyyy-mm-dd", 13), ("报告期在任", lambda r, _: "是" if r.get("active_on_report_date") else "否", None, 12),
        ("成立日期", "inception_date", "yyyy-mm-dd", 13), ("报告期核实经理", "verified_manager", None, 16),
        ("核验状态", "manager_verification", None, 18), ("是否纳入", lambda r, _: "是" if r.get("selected") else "否", None, 10),
        ("筛选结论", "selection_reason", None, 34), ("经理档案URL", "manager_profile_url", None, 24),
        ("基金信息URL", "fund_info_url", None, 24), ("经理历史URL", "manager_history_url", None, 24),
    ]
    for name, title, rows in (
        ("报告期基金池", "报告期基金池与产品筛选", data.get("selected_funds", [])),
        ("全部任职历史", "基金经理完整任职历史", data.get("all_tenures", [])),
    ):
        ws = add_table_sheet(workbook, name, title, _headers(columns), _rows(rows, columns), widths=_widths(columns), title_row=False)
        _apply_column_formats(ws, columns, 2, ws.max_row)
    issue_columns: list[Column] = [
        ("级别", "severity", None, 10), ("分类", "category", None, 18), ("基金代码", "fund_code", "@", 12),
        ("基金名称", "fund_name", None, 24), ("基金经理", "manager", None, 14), ("报告期", "report_date", "yyyy-mm-dd", 13),
        ("问题说明", "message", None, 40), ("来源URL", "source_url", None, 28), ("建议处理", "action", None, 32),
    ]
    add_table_sheet(workbook, "核验与异常", "核验与异常", _headers(issue_columns), _rows(data.get("issues", []), issue_columns), widths=_widths(issue_columns), title_row=False)
    source_rows = [[r.get("source_id", ""), r.get("name", ""), r.get("purpose", ""), r.get("url", "")] for r in data.get("sources", [])]
    source_rows.extend([
        ["口径", "报告期基金池规则", "任职开始≤报告期≤任职结束；成立日晚于报告期时排除", ""],
        ["去重", "份额处理规则", "本步骤保留全部A/C/E份额，持仓阶段再确定正式代表份额", ""],
    ])
    add_table_sheet(workbook, "来源与口径", "来源与口径", ["来源ID", "来源名称", "用途", "URL"], source_rows, title_row=False)
    add_summary_sheet(
        workbook, "基金经理报告期基金池运行摘要",
        [("基金经理", summary.get("manager")), ("天天基金经理ID", summary.get("manager_id")), ("基金公司", summary.get("company")),
         ("报告期", summary.get("report_date")), ("历史任职份额数", summary.get("historical_share_count", 0)),
         ("报告期在任份额数", summary.get("active_share_count", 0)), ("最终纳入份额数", summary.get("selected_share_count", 0)),
         ("基础产品数（未去重）", summary.get("product_count", 0)), ("经理历史核验通过", summary.get("verified_count", 0)),
         ("运行状态", "通过" if not summary.get("error_count") else "需核查")],
        [("错误", summary.get("error_count", 0)), ("警告", summary.get("warning_count", 0)), ("异常合计", len(data.get("issues", [])))],
        note="口径：以基金经理档案中的完整历史任职区间建立报告期基金池，再用单只基金基本信息与经理变动历史交叉核验。A/C/E等份额本表全部保留。",
    )
    _move_summary_first(workbook)
    return _save(workbook, output_path)


def _holding_columns(include_manager: bool = True, include_industry: bool = False) -> list[Column]:
    columns: list[Column] = [("基金代码", "fund_code", "@", 12), ("基金名称", "fund_name", None, 24)]
    if include_manager:
        columns.append(("基金经理", "manager", None, 14))
    columns.extend([
        ("报告期", "report_date", "yyyy-mm-dd", 13), ("序号", "rank", "0", 8), ("股票代码", "stock_code", "@", 13),
        ("股票名称", "stock_name", None, 16), ("持股数量(万股)", "shares_10k", "#,##0.00", 16),
        ("持仓市值(万元)", "market_value_10k", "#,##0.00", 17), ("占基金净值比例", "nav_ratio", "0.00%", 16),
        ("市场/地区", "market", None, 12), ("重复组", "duplicate_group", None, 14),
        ("代表份额", lambda r, _: "是" if r.get("representative") else "否", None, 10), ("来源URL", "source_url", None, 28),
    ])
    if include_industry:
        columns = [
            ("基金代码", "fund_code", "@", 12), ("基金名称", "fund_name", None, 24), ("报告期", "report_date", "yyyy-mm-dd", 13),
            ("序号", "rank", "0", 8), ("股票代码", "stock_code", "@", 13), ("股票名称", "stock_name", None, 16),
            ("市场/地区", "market", None, 12), ("持仓市值(万元)", "market_value_10k", "#,##0.00", 17),
            ("占基金净值比例", "nav_ratio", "0.00%", 16), ("申万一级行业", "sw_level1", None, 16),
            ("申万二级行业", "sw_level2", None, 18), ("分类快照日期", "industry_snapshot_date", "yyyy-mm-dd", 14),
            ("行业来源ID", "industry_source_id", None, 14),
        ]
    return columns


def build_holdings_report(input_path: Path, output_path: Path) -> Path:
    data = read_json(input_path)
    summary = data["summary"]
    quality = data.get("industry_quality", {})
    workbook = _new_workbook()
    columns = _holding_columns()
    add_table_sheet(workbook, "正式版_持仓明细", "正式版持仓明细", _headers(columns), _rows(data.get("formal_holdings", []), columns), widths=_widths(columns), title_row=False)
    add_table_sheet(workbook, "全量抓取底稿", "全量抓取底稿", _headers(columns), _rows(data.get("all_holdings", []), columns), widths=_widths(columns), title_row=False)
    fund_columns: list[Column] = [
        ("名单基金经理", "manager", None, 14), ("基金代码", "fund_code", "@", 12), ("基金名称", "fund_name", None, 24),
        ("基金类型", "fund_type", None, 18), ("成立日期", "inception_date", "yyyy-mm-dd", 13),
        ("是否纳入", lambda r, _: "是" if r.get("selected") else "否", None, 10), ("筛选原因", "selection_reason", None, 32),
        ("报告期核实经理", "verified_manager", None, 16), ("经理核验状态", "manager_status", None, 16),
        ("抓取状态", "fetch_status", None, 16), ("经理来源URL", "manager_source_url", None, 28),
    ]
    add_table_sheet(workbook, "基金名单与筛选", "基金名单与筛选", _headers(fund_columns), _rows(data.get("funds", []), fund_columns), widths=_widths(fund_columns), title_row=False)
    issue_columns: list[Column] = [
        ("级别", "severity", None, 10), ("分类", "category", None, 18), ("基金代码", "fund_code", "@", 12),
        ("基金名称", "fund_name", None, 24), ("名单基金经理", "manager", None, 14), ("报告期", "report_date", "yyyy-mm-dd", 13),
        ("问题说明", "message", None, 40), ("来源URL", "source_url", None, 28), ("建议处理", "action", None, 32),
    ]
    add_table_sheet(workbook, "异常清单", "异常清单", _headers(issue_columns), _rows(data.get("issues", []), issue_columns), widths=_widths(issue_columns), title_row=False)
    industry_columns = _holding_columns(include_manager=False, include_industry=True)
    add_table_sheet(workbook, "正式持仓与行业", "正式持仓与申万行业", _headers(industry_columns), _rows(data.get("formal_holdings_industry", []), industry_columns), widths=_widths(industry_columns), title_row=False)
    summary_rows = []
    holding_end = max(2, len(data.get("formal_holdings_industry", [])) + 1)
    for item_index, row in enumerate(data.get("industry_summary", []), start=2):
        summary_rows.append([
            row.get("fund_code", ""), row.get("fund_name", ""), row.get("sw_level1", ""),
            f'=COUNTIFS(\'正式持仓与行业\'!$A$2:$A${holding_end},A{item_index},\'正式持仓与行业\'!$J$2:$J${holding_end},C{item_index})',
            f'=SUMIFS(\'正式持仓与行业\'!$H$2:$H${holding_end},\'正式持仓与行业\'!$A$2:$A${holding_end},A{item_index},\'正式持仓与行业\'!$J$2:$J${holding_end},C{item_index})',
            f'=SUMIFS(\'正式持仓与行业\'!$I$2:$I${holding_end},\'正式持仓与行业\'!$A$2:$A${holding_end},A{item_index},\'正式持仓与行业\'!$J$2:$J${holding_end},C{item_index})',
            f'=IFERROR(E{item_index}/SUMIF($A$2:$A${len(data.get("industry_summary", []))+1},A{item_index},$E$2:$E${len(data.get("industry_summary", []))+1}),0)',
        ])
    add_table_sheet(workbook, "基金行业汇总", "基金行业汇总", ["基金代码", "基金名称", "申万一级行业", "重仓股数量", "持仓市值合计(万元)", "净值比例合计", "前十持仓市值占比"], summary_rows, title_row=False)
    mapping_columns: list[Column] = [
        ("行业来源ID", "industry_source_id", None, 14), ("股票代码", "stock_code", "@", 13), ("股票名称", "stock_name", None, 16),
        ("市场/地区", "market", None, 12), ("申万一级行业", "sw_level1", None, 16), ("申万二级行业", "sw_level2", None, 18),
        ("申万二级代码", "sw_level2_code", "@", 14), ("分类快照日期", "industry_snapshot_date", "yyyy-mm-dd", 14),
        ("分类状态", "classification_status", None, 18), ("来源URL", "source_url", None, 28),
    ]
    add_table_sheet(workbook, "股票行业映射", "股票行业映射", _headers(mapping_columns), _rows(data.get("stock_industry_mapping", []), mapping_columns), widths=_widths(mapping_columns), title_row=False)
    industry_issue_columns: list[Column] = [
        ("级别", "severity", None, 10), ("分类", "category", None, 20), ("股票代码", "stock_code", "@", 13),
        ("股票名称", "stock_name", None, 16), ("问题说明", "message", None, 45), ("来源URL", "source_url", None, 28), ("建议处理", "action", None, 34),
    ]
    add_table_sheet(workbook, "行业异常", "行业异常", _headers(industry_issue_columns), _rows(data.get("industry_issues", []), industry_issue_columns), widths=_widths(industry_issue_columns), title_row=False)
    quality_items = [
        ("正式持仓行业覆盖", quality.get("mapped_holding_rows", 0), quality.get("eligible_holding_rows", 0), "A股正式持仓应匹配申万行业"),
        ("唯一股票行业覆盖", quality.get("unique_stock_mapped", 0), quality.get("unique_stock_count", 0), "唯一股票代码不得因重复持仓重复计算"),
        ("持仓行业错误", quality.get("error_count", 0), 0, "行业错误应为0"),
        ("历史行业时点", "当前快照" if quality.get("historical_point_in_time") else "历史时点", "历史时点", "当前公开快照限制必须显式披露"),
    ]
    quality_rows = []
    for index, (item, actual, expected, note) in enumerate(quality_items, start=2):
        quality_rows.append([item, actual, expected, f'=IF(AND(ISNUMBER(B{index}),ISNUMBER(C{index})),B{index}-C{index},"")', f'=IF(B{index}=C{index},"OK","CHECK")', note])
    quality_ws = add_table_sheet(workbook, "行业数据质量", "行业数据质量", ["检查项", "实际值", "预期值", "差异", "状态", "说明"], quality_rows, title_row=False)
    quality_ws.conditional_formatting.add(f"E2:E{quality_ws.max_row}", FormulaRule(formula=['E2="OK"'], fill=PatternFill("solid", fgColor=GREEN)))
    source_rows = [[r.get("item", ""), r.get("note", ""), r.get("url", "")] for r in data.get("sources", []) + data.get("industry_sources", [])]
    source_rows.extend([
        ["报告期", summary.get("report_date", ""), ""],
        ["持仓口径", "正式版按A/C/E去重代表份额；全量份额保留在抓取底稿", ""],
        ["行业口径", "A股采用申万一级当前公开快照；港股及海外不强行映射", ""],
    ])
    add_table_sheet(workbook, "来源与口径", "来源与口径", ["项目", "说明", "来源URL"], source_rows, title_row=False)
    add_summary_sheet(
        workbook, "基金季度前十大重仓运行摘要",
        [("报告期", summary.get("report_date")), ("输入基金", summary.get("input_funds", 0)), ("纳入基金份额", summary.get("selected_funds", 0)),
         ("成功基金份额", summary.get("successful_funds", 0)), ("正式代表基金", summary.get("formal_funds", 0)),
         ("正式持仓行数", summary.get("formal_holding_rows", 0)), ("运行状态", "通过" if not summary.get("error_count") else "需核查")],
        [("管道错误", summary.get("error_count", 0)), ("管道警告", summary.get("warning_count", 0)),
         ("A股行业覆盖率", quality.get("holding_coverage", 0)), ("行业错误", quality.get("error_count", 0)), ("行业警告", quality.get("warning_count", 0))],
        note="正式版保留A/C/E去重后的代表份额前十大持仓；行业为公开可得申万当前快照，历史报告期需保留时点限制。",
    )
    _move_summary_first(workbook)
    return _save(workbook, output_path)


def build_resource_report(input_path: Path, output_path: Path) -> Path:
    data = read_json(input_path)
    summary = data["summary"]
    workbook = _new_workbook()
    industry_columns: list[Column] = [
        ("优先级", "priority", None, 10), ("申万一级行业", "sw_level1", None, 16), ("涉及基金数", "fund_count", "0", 12),
        ("基金代码", lambda r, _: "、".join(r.get("fund_codes", [])), None, 26), ("重仓记录数", "holding_count", "0", 12),
        ("持仓市值合计(万元)", "market_value_10k", "#,##0.00", 18), ("净值比例算术合计", "nav_ratio_sum", "0.00%", 17),
        ("匹配状态", "match_status", None, 16),
    ]
    add_table_sheet(workbook, "行业对接需求", "行业对接需求", _headers(industry_columns), _rows(data.get("industry_demands", []), industry_columns), widths=_widths(industry_columns), title_row=False)
    company_columns: list[Column] = [
        ("优先级", "priority", None, 10), ("股票代码", "stock_code", "@", 13), ("股票名称", "stock_name", None, 16),
        ("申万一级行业", "sw_level1", None, 16), ("涉及基金数", "fund_count", "0", 12),
        ("基金代码", lambda r, _: "、".join(r.get("fund_codes", [])), None, 28), ("持仓出现次数", "holding_occurrences", "0", 13),
        ("持仓市值合计(万元)", "market_value_10k", "#,##0.00", 18), ("单基金最高净值比例", "max_nav_ratio", "0.00%", 18),
        ("匹配状态", "match_status", None, 16),
    ]
    add_table_sheet(workbook, "公司对接需求", "公司对接需求", _headers(company_columns), _rows(data.get("company_demands", []), company_columns), widths=_widths(company_columns), title_row=False)
    match_columns: list[Column] = [
        ("需求类型", "demand_type", None, 12), ("优先级", "priority", None, 10), ("目标代码", "target_code", "@", 14),
        ("目标名称", "target_name", None, 18), ("申万一级行业", "sw_level1", None, 16), ("匹配方式", "match_type", None, 20),
        ("匹配分", "score", "0", 10), ("人员姓名", "person_name", None, 14), ("所属机构", "organization", None, 22),
        ("人员类型", "person_type", None, 12), ("专长标签", "expertise_tags", None, 24), ("地区", "region", None, 12),
        ("联系权限", "contact_permission", None, 12), ("联系方式", "contact_info", None, 24),
    ]
    add_table_sheet(workbook, "匹配结果", "匹配结果", _headers(match_columns), _rows(data.get("matches", []), match_columns), widths=_widths(match_columns), title_row=False)
    pending_columns: list[Column] = [
        ("优先级", "priority", None, 10), ("需求类型", "demand_type", None, 12), ("目标代码", "target_code", "@", 14),
        ("目标名称", "target_name", None, 18), ("未匹配原因", "reason", None, 34), ("建议处理", "action", None, 34),
    ]
    add_table_sheet(workbook, "待补充项", "待补充项", _headers(pending_columns), _rows(data.get("pending_items", []), pending_columns), widths=_widths(pending_columns), title_row=False)
    personnel_columns: list[Column] = [
        ("人员姓名", "person_name", None, 14), ("所属机构", "organization", None, 22), ("人员类型", "person_type", None, 12),
        ("覆盖申万一级行业", "sw_level1", None, 24), ("覆盖公司代码", "covered_stock_codes", None, 24),
        ("专长标签", "expertise_tags", None, 24), ("地区", "region", None, 12), ("当前状态", "current_status", None, 12),
        ("联系权限", "contact_permission", None, 12), ("联系方式", "contact_info", None, 24),
    ]
    add_table_sheet(workbook, "人员库模板", "人员库模板", _headers(personnel_columns), _rows(data.get("personnel_rows", []), personnel_columns), widths=_widths(personnel_columns), title_row=False)
    issue_columns: list[Column] = [
        ("级别", "severity", None, 10), ("分类", "category", None, 18), ("原始行号", "row_number", "0", 12),
        ("人员姓名", "person_name", None, 14), ("所属机构", "organization", None, 22), ("问题说明", "message", None, 36), ("建议处理", "action", None, 32),
    ]
    add_table_sheet(workbook, "人员库校验", "人员库校验", _headers(issue_columns), _rows(data.get("personnel_issues", []), issue_columns), widths=_widths(issue_columns), title_row=False)
    add_table_sheet(workbook, "匹配口径", "匹配口径", ["项目", "确定性规则"], [[r.get("item", ""), r.get("rule", "")] for r in data.get("rules", [])], widths={"确定性规则": 75}, title_row=False)
    add_summary_sheet(
        workbook, "基金重仓研究资源对接准备",
        [("报告期", summary.get("report_date")), ("人员库人数", summary.get("personnel_count", 0)), ("在岗人员", summary.get("active_personnel_count", 0)),
         ("行业需求", summary.get("industry_demand_count", 0)), ("公司需求", summary.get("company_demand_count", 0)),
         ("匹配记录", summary.get("match_count", 0)), ("待补充项", summary.get("pending_count", 0)), ("运行状态", summary.get("status", ""))],
        [("人员库错误", summary.get("personnel_error_count", 0)), ("模型调用", "未调用 DeepSeek")],
        note="确定性匹配优先使用公司精确覆盖和行业覆盖；联系方式只在权限允许时展示，未确认内容进入待补充或候选清单。",
        right_title="优先处理指标",
    )
    _move_summary_first(workbook)
    return _save(workbook, output_path)


def build_quarter_comparison_report(input_path: Path, output_path: Path) -> Path:
    data = read_json(input_path)
    summary = data["summary"]
    workbook = _new_workbook()
    company_columns: list[Column] = [
        ("变化类型", lambda _r, i: f'=IF(E{i}=0,"新进",IF(F{i}=0,"退出",IF(I{i}>0.01,"增持",IF(I{i}<-0.01,"减持","持平"))))', None, 12),
        ("股票代码", "stock_code", "@", 13), ("股票名称", "stock_name", None, 16), ("申万一级行业", "sw_level1", None, 16),
        ("上期持有基金数", "previous_fund_count", "0", 14), ("本期持有基金数", "current_fund_count", "0", 14),
        ("上期持股(万股)", "previous_shares_10k", "#,##0.00", 15), ("本期持股(万股)", "current_shares_10k", "#,##0.00", 15),
        ("持股变化(万股)", lambda _r, i: f'=H{i}-G{i}', "#,##0.00", 15), ("持股变化率", lambda _r, i: f'=IFERROR(I{i}/G{i},"")', "0.00%", 14),
        ("上期市值(万元)", "previous_market_value_10k", "#,##0.00", 16), ("本期市值(万元)", "current_market_value_10k", "#,##0.00", 16),
        ("市值变化(万元)", lambda _r, i: f'=L{i}-K{i}', "#,##0.00", 16), ("市值变化率", lambda _r, i: f'=IFERROR(M{i}/K{i},"")', "0.00%", 14),
        ("上期净值比例合计", "previous_nav_ratio_sum", "0.00%", 16), ("本期净值比例合计", "current_nav_ratio_sum", "0.00%", 16),
        ("净值比例变化", lambda _r, i: f'=P{i}-O{i}', "0.00%", 15), ("上期最佳排名", "previous_best_rank", "0", 13),
        ("本期最佳排名", "current_best_rank", "0", 13), ("排名提升", lambda _r, i: f'=IF(OR(R{i}="",S{i}=""),"",R{i}-S{i})', "0", 12),
        ("上期基金代码", lambda r, _: "、".join(r.get("previous_fund_codes", [])), None, 24), ("本期基金代码", lambda r, _: "、".join(r.get("current_fund_codes", [])), None, 24),
    ]
    add_table_sheet(workbook, "公司变化汇总", "公司变化汇总", _headers(company_columns), _rows(data.get("company_changes", []), company_columns), widths=_widths(company_columns), title_row=False)
    fund_columns: list[Column] = [
        ("基金代码", "fund_code", "@", 12), ("基金名称", "fund_name", None, 24),
        ("变化类型", lambda _r, i: f'=IF(G{i}="","新进",IF(H{i}="","退出",IF(L{i}>0.01,"增持",IF(L{i}<-0.01,"减持","持平"))))', None, 12),
        ("股票代码", "stock_code", "@", 13), ("股票名称", "stock_name", None, 16), ("申万一级行业", "sw_level1", None, 16),
        ("上期排名", "previous_rank", "0", 11), ("本期排名", "current_rank", "0", 11),
        ("排名提升", lambda _r, i: f'=IF(OR(G{i}="",H{i}=""),"",G{i}-H{i})', "0", 11),
        ("上期持股(万股)", "previous_shares_10k", "#,##0.00", 15), ("本期持股(万股)", "current_shares_10k", "#,##0.00", 15),
        ("持股变化(万股)", lambda _r, i: f'=K{i}-J{i}', "#,##0.00", 15), ("持股变化率", lambda _r, i: f'=IFERROR(L{i}/J{i},"")', "0.00%", 14),
        ("上期市值(万元)", "previous_market_value_10k", "#,##0.00", 16), ("本期市值(万元)", "current_market_value_10k", "#,##0.00", 16),
        ("市值变化(万元)", lambda _r, i: f'=O{i}-N{i}', "#,##0.00", 16), ("上期净值比例", "previous_nav_ratio", "0.00%", 14),
        ("本期净值比例", "current_nav_ratio", "0.00%", 14), ("净值比例变化", lambda _r, i: f'=R{i}-Q{i}', "0.00%", 14),
    ]
    add_table_sheet(workbook, "基金内持仓变化", "基金内持仓变化", _headers(fund_columns), _rows(data.get("fund_stock_changes", []), fund_columns), widths=_widths(fund_columns), title_row=False)
    industry_columns: list[Column] = [
        ("变化类型", lambda _r, i: f'=IF(E{i}=0,"新进入前十行业",IF(F{i}=0,"退出前十行业",IF(M{i}>0.0001,"上升",IF(M{i}<-0.0001,"下降","持平"))))', None, 18),
        ("申万一级行业", "sw_level1", None, 16), ("上期涉及基金数", "previous_fund_count", "0", 14), ("本期涉及基金数", "current_fund_count", "0", 14),
        ("上期重仓记录数", "previous_holding_count", "0", 14), ("本期重仓记录数", "current_holding_count", "0", 14),
        ("记录数变化", lambda _r, i: f'=F{i}-E{i}', "0", 12), ("上期市值(万元)", "previous_market_value_10k", "#,##0.00", 16),
        ("本期市值(万元)", "current_market_value_10k", "#,##0.00", 16), ("市值变化(万元)", lambda _r, i: f'=I{i}-H{i}', "#,##0.00", 16),
        ("上期净值比例合计", "previous_nav_ratio_sum", "0.00%", 16), ("本期净值比例合计", "current_nav_ratio_sum", "0.00%", 16),
        ("净值比例变化", lambda _r, i: f'=L{i}-K{i}', "0.00%", 15),
        ("上期基金代码", lambda r, _: "、".join(r.get("previous_fund_codes", [])), None, 24), ("本期基金代码", lambda r, _: "、".join(r.get("current_fund_codes", [])), None, 24),
    ]
    add_table_sheet(workbook, "行业变化", "行业变化", _headers(industry_columns), _rows(data.get("industry_changes", []), industry_columns), widths=_widths(industry_columns), title_row=False)
    add_table_sheet(workbook, "数据质量", "数据质量", ["检查项", "实际值", "预期值", "状态", "说明"], [[r.get("item"), r.get("actual"), r.get("expected"), r.get("status"), r.get("note")] for r in data.get("checks", [])], title_row=False)
    add_table_sheet(workbook, "比较口径", "比较口径", ["项目", "确定性规则"], [[r.get("item"), r.get("rule")] for r in data.get("rules", [])], widths={"确定性规则": 80}, title_row=False)
    add_table_sheet(workbook, "来源与审计", "来源与审计", ["来源项目", "报告期", "本地审计文件", "公开来源"], [[r.get("item"), r.get("report_date"), r.get("path"), "天天基金/东方财富公开披露"] for r in data.get("sources", [])], widths={"本地审计文件": 55}, title_row=False)
    add_summary_sheet(
        workbook, "基金季度前十大持仓变化分析",
        [("基金经理", summary.get("manager")), ("上期报告日", summary.get("previous_report_date")), ("本期报告日", summary.get("current_report_date")),
         ("上期代表基金", summary.get("previous_formal_funds", 0)), ("本期代表基金", summary.get("current_formal_funds", 0)),
         ("上期正式持仓行", summary.get("previous_holding_rows", 0)), ("本期正式持仓行", summary.get("current_holding_rows", 0)), ("运行状态", summary.get("status", ""))],
        [("公司并集", summary.get("company_union_count", 0)), ("新进", summary.get("new_company_count", 0)), ("退出", summary.get("exited_company_count", 0)),
         ("增持", summary.get("increased_company_count", 0)), ("减持", summary.get("decreased_company_count", 0)),
         ("行业并集", summary.get("industry_union_count", 0)), ("行业快照", summary.get("industry_snapshot_date", ""))],
        note="变化基于两期公开披露的前十大持仓，不代表完整交易流水。行业比较使用同一公开快照，历史时点限制必须保留。",
    )
    _move_summary_first(workbook)
    return _save(workbook, output_path)


def _company_label(value: str) -> str:
    text = value
    for suffix in ("基金管理股份有限公司", "基金管理有限公司", "基金有限责任公司", "管理有限公司", "有限公司"):
        if text.endswith(suffix):
            text = text[: -len(suffix)]
            break
    return text + ("基金" if not text.endswith("基金") else "")


def _load_manager_artifacts(summary: dict[str, Any]) -> list[dict[str, Any]]:
    result = []
    for manager in summary.get("manager_results", []):
        root = Path(str(manager.get("output_dir", "")))
        payload: dict[str, Any] = {"result": manager, "root": root}
        for key, filename in (
            ("pool", "manager_fund_pool_data.json"), ("pipeline", "pipeline_data.json"),
            ("industry", "industry_analysis_data.json"), ("readiness", "disclosure_readiness.json"),
            ("manifest", "batch_manifest.json"),
        ):
            path = root / filename
            payload[key] = read_json(path) if path.exists() else {}
        result.append(payload)
    return result


def _company_portfolio_rows(summary: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    managers = _load_manager_artifacts(summary)
    overview: list[dict[str, Any]] = []
    fund_pool: list[dict[str, Any]] = []
    manager_holdings: list[dict[str, Any]] = []
    anomalies: list[dict[str, Any]] = []
    for item in managers:
        result = item["result"]
        pool = item["pool"]
        pipeline = item["pipeline"]
        industry = item["industry"]
        readiness = item["readiness"]
        manager = str(result.get("manager", ""))
        pool_summary = pool.get("summary", {})
        pipeline_summary = pipeline.get("summary", {})
        quality = industry.get("industry_quality", {})
        formal = industry.get("formal_holdings_industry", [])
        a_stocks = {row.get("stock_code") for row in formal if row.get("market") == "A股"}
        mapped = {row.get("stock_code") for row in formal if row.get("market") == "A股" and row.get("industry_status") == "当前快照已匹配"}
        overview.append({
            "manager": manager, "manager_id": result.get("manager_id", ""),
            "analysis_status": "有适用产品" if pool_summary.get("selected_share_count", 0) else "无适用产品",
            "active_share_count": pool_summary.get("active_share_count", 0), "selected_share_count": pool_summary.get("selected_share_count", 0),
            "product_count": pool_summary.get("product_count", 0), "successful_funds": pipeline_summary.get("successful_funds", 0),
            "formal_funds": pipeline_summary.get("formal_funds", 0), "formal_holding_rows": pipeline_summary.get("formal_holding_rows", 0),
            "error_count": pipeline_summary.get("error_count", 0), "warning_count": pipeline_summary.get("warning_count", 0),
            "unique_stock_count": len({row.get("stock_code") for row in formal}), "a_coverage": len(mapped) / len(a_stocks) if a_stocks else 1.0,
            "workflow_status": result.get("overall_status", ""), "readiness_status": readiness.get("summary", {}).get("status", ""),
        })
        for row in pool.get("all_tenures", []):
            if row.get("active_on_report_date"):
                fund_pool.append({"target_manager": manager, **row})
            if row.get("active_on_report_date") and not row.get("selected"):
                anomalies.append({
                    "manager": manager, "type": "产品排除", "severity": "提示", "category": "基金筛选",
                    "code": row.get("fund_code", ""), "name": row.get("fund_name", ""), "report_date": summary.get("report_date", ""),
                    "message": row.get("selection_reason", ""), "blocking": "否", "action": "保留审计，不进入正式持仓", "source_url": row.get("fund_info_url", ""),
                })
        for row in formal:
            manager_holdings.append({"target_manager": manager, **row, "verified_manager": row.get("manager", manager)})
        for row in pipeline.get("issues", []):
            anomalies.append({
                "manager": manager, "type": "持仓数据", "severity": row.get("severity", ""), "category": row.get("category", ""),
                "code": row.get("fund_code", ""), "name": row.get("fund_name", ""), "report_date": row.get("report_date", summary.get("report_date", "")),
                "message": row.get("message", ""), "blocking": "是" if row.get("severity") == "错误" else "否", "action": row.get("action", ""), "source_url": row.get("source_url", ""),
            })
        for row in industry.get("industry_issues", []):
            anomalies.append({
                "manager": manager, "type": "行业数据", "severity": row.get("severity", ""), "category": row.get("category", ""),
                "code": row.get("stock_code", ""), "name": row.get("stock_name", ""), "report_date": summary.get("report_date", ""),
                "message": row.get("message", ""), "blocking": "是" if row.get("severity") == "错误" else "否", "action": row.get("action", ""), "source_url": row.get("source_url", ""),
            })

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in manager_holdings:
        grouped[(str(row.get("fund_code", "")), str(row.get("stock_code", "")))].append(row)
    company_holdings: list[dict[str, Any]] = []
    for group in grouped.values():
        base = dict(group[0])
        base["managers"] = "、".join(sorted({str(row.get("target_manager", "")) for row in group if row.get("target_manager")}))
        company_holdings.append(base)
    company_holdings.sort(key=lambda row: (str(row.get("fund_code", "")), int(row.get("rank", 999)), str(row.get("stock_code", ""))))

    stock_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    industry_groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in company_holdings:
        stock_groups[str(row.get("stock_code", ""))].append(row)
        label = str(row.get("sw_level1") or row.get("market") or "未分类")
        industry_groups[(label, str(row.get("market", "")))].append(row)
    stock_rollup = []
    for code, group in stock_groups.items():
        first = group[0]
        stock_rollup.append({
            "stock_code": code, "market": first.get("market", ""), "stock_name": first.get("stock_name", ""),
            "fund_count": len({r.get("fund_code") for r in group}), "manager_count": len({m for r in group for m in str(r.get("managers", "")).split("、") if m}),
            "managers": "、".join(sorted({m for r in group for m in str(r.get("managers", "")).split("、") if m})),
            "record_count": len(group), "shares_sum": sum(float(r.get("shares_10k") or 0) for r in group),
            "market_value_sum": sum(float(r.get("market_value_10k") or 0) for r in group), "nav_ratio_sum": sum(float(r.get("nav_ratio") or 0) for r in group),
            "industry_status": first.get("industry_status", ""), "sw_level1": first.get("sw_level1", ""), "sw_level2": first.get("sw_level2", ""),
            "industry_source_url": first.get("industry_source_url", ""),
        })
    stock_rollup.sort(key=lambda row: (-row["fund_count"], -row["market_value_sum"], row["stock_code"]))
    industry_rollup = []
    for (label, market), group in industry_groups.items():
        industry_rollup.append({
            "label": label, "market": market, "stock_count": len({r.get("stock_code") for r in group}),
            "fund_count": len({r.get("fund_code") for r in group}), "manager_count": len({m for r in group for m in str(r.get("managers", "")).split("、") if m}),
            "record_count": len(group), "market_value_sum": sum(float(r.get("market_value_10k") or 0) for r in group),
            "nav_ratio_sum": sum(float(r.get("nav_ratio") or 0) for r in group),
            "snapshot_date": next((r.get("industry_snapshot_date", "") for r in group if r.get("industry_snapshot_date")), ""),
            "note": "申万一级当前快照" if market == "A股" else "非A股不强制使用申万分类",
        })
    industry_rollup.sort(key=lambda row: (-row["market_value_sum"], row["label"]))
    return {
        "overview": overview, "fund_pool": fund_pool, "manager_holdings": manager_holdings,
        "company_holdings": company_holdings, "stock_rollup": stock_rollup, "industry_rollup": industry_rollup,
        "anomalies": anomalies,
    }


def build_company_portfolio_report(input_path: Path, output_path: Path) -> Path:
    data = read_json(input_path)
    rows = _company_portfolio_rows(data)
    workbook = _new_workbook()
    company = str((data.get("companies") or ["基金公司"])[0])
    label = _company_label(company)
    report_date = str(data.get("report_date", ""))
    quarter = f"{report_date[:4]}Q{(int(report_date[5:7]) - 1) // 3 + 1}" if report_date else ""
    overview_columns: list[Column] = [
        ("目标基金经理", "manager", None, 14), ("天天基金经理ID", "manager_id", "@", 16), ("分析状态", "analysis_status", None, 14),
        ("报告期在任份额", "active_share_count", "0", 14), ("纳入份额", "selected_share_count", "0", 11), ("经理-产品数", "product_count", "0", 13),
        ("成功份额", "successful_funds", "0", 11), ("正式代表产品", "formal_funds", "0", 14), ("正式持仓行数", "formal_holding_rows", "0", 14),
        ("持仓错误", "error_count", "0", 11), ("持仓警告", "warning_count", "0", 11), ("唯一重仓股", "unique_stock_count", "0", 12),
        ("A股行业覆盖率", "a_coverage", "0.00%", 14), ("工作流状态", "workflow_status", None, 14), ("披露状态", "readiness_status", None, 16),
    ]
    add_table_sheet(workbook, "基金经理概览", f"{label} {quarter} 基金经理覆盖与运行状态", _headers(overview_columns), _rows(rows["overview"], overview_columns), widths=_widths(overview_columns))
    pool_columns: list[Column] = [
        ("目标基金经理", "target_manager", None, 14), ("基金代码", "fund_code", "@", 12), ("基金名称", "fund_name", None, 24),
        ("基础产品名称", "product_base_name", None, 22), ("产品组", "product_group", None, 14), ("基金类型", "fund_type", None, 18),
        ("任职开始", "tenure_start", "yyyy-mm-dd", 13), ("任职结束", "tenure_end", "yyyy-mm-dd", 13), ("成立日期", "inception_date", "yyyy-mm-dd", 13),
        ("报告期核实经理", "verified_manager", None, 16), ("核验状态", "manager_verification", None, 16),
        ("是否纳入", lambda r, _: "是" if r.get("selected") else "否", None, 10), ("筛选结论", "selection_reason", None, 34),
        ("经理档案URL", "manager_profile_url", None, 26), ("基金信息URL", "fund_info_url", None, 26), ("经理历史URL", "manager_history_url", None, 26),
    ]
    add_table_sheet(workbook, "基金池明细", "报告期基金池、产品筛选与经理核验", _headers(pool_columns), _rows(rows["fund_pool"], pool_columns), widths=_widths(pool_columns))
    manager_holding_columns: list[Column] = [
        ("目标基金经理", "target_manager", None, 14), ("基金代码", "fund_code", "@", 12), ("基金名称", "fund_name", None, 24),
        ("报告期核实经理", "verified_manager", None, 16), ("报告期", "report_date", "yyyy-mm-dd", 13), ("排名", "rank", "0", 8),
        ("股票代码", "stock_code", "@", 13), ("股票名称", "stock_name", None, 16), ("市场", "market", None, 10),
        ("持股数量(万股)", "shares_10k", "#,##0.00", 15), ("持仓市值(万元)", "market_value_10k", "#,##0.00", 16),
        ("占基金净值比例", "nav_ratio", "0.00%", 15), ("申万一级行业", "sw_level1", None, 15), ("申万二级行业", "sw_level2", None, 17),
        ("行业匹配状态", "industry_status", None, 18), ("重复组", "duplicate_group", None, 14),
        ("代表份额", lambda r, _: "是" if r.get("representative") else "否", None, 10), ("持仓来源URL", "source_url", None, 26),
        ("行业来源URL", "industry_source_url", None, 26),
    ]
    add_table_sheet(workbook, "经理持仓明细", "经理口径正式持仓（经理－代表产品）", _headers(manager_holding_columns), _rows(rows["manager_holdings"], manager_holding_columns), widths=_widths(manager_holding_columns))
    company_holding_columns: list[Column] = [
        ("基金代码", "fund_code", "@", 12), ("基金名称", "fund_name", None, 24), ("本名单涉及经理", "managers", None, 24),
        ("报告期核实经理", "verified_manager", None, 16), ("报告期", "report_date", "yyyy-mm-dd", 13), ("排名", "rank", "0", 8),
        ("股票代码", "stock_code", "@", 13), ("股票名称", "stock_name", None, 16), ("市场", "market", None, 10),
        ("持股数量(万股)", "shares_10k", "#,##0.00", 15), ("持仓市值(万元)", "market_value_10k", "#,##0.00", 16),
        ("占基金净值比例", "nav_ratio", "0.00%", 15), ("申万一级行业", "sw_level1", None, 15), ("申万二级行业", "sw_level2", None, 17),
        ("行业匹配状态", "industry_status", None, 18), ("持仓来源URL", "source_url", None, 26), ("行业来源URL", "industry_source_url", None, 26),
    ]
    add_table_sheet(workbook, "公司唯一基金持仓", "公司口径持仓（跨经理按基金代码＋股票代码去重）", _headers(company_holding_columns), _rows(rows["company_holdings"], company_holding_columns), widths=_widths(company_holding_columns))
    stock_columns: list[Column] = [
        ("股票代码", "stock_code", "@", 13), ("市场", "market", None, 10), ("股票名称", "stock_name", None, 16),
        ("涉及基金数", "fund_count", "0", 12), ("涉及经理数", "manager_count", "0", 12), ("本名单涉及经理", "managers", None, 26),
        ("基金-股票记录数", "record_count", "0", 15), ("持股数量合计(万股)", "shares_sum", "#,##0.00", 18),
        ("持仓市值合计(万元)", "market_value_sum", "#,##0.00", 19), ("净值比例合计(非加权)", "nav_ratio_sum", "0.00%", 19),
        ("行业匹配状态", "industry_status", None, 18), ("申万一级行业", "sw_level1", None, 15), ("申万二级行业", "sw_level2", None, 17),
        ("行业来源URL", "industry_source_url", None, 28),
    ]
    add_table_sheet(workbook, "重仓公司汇总", "公司口径重仓股票汇总（唯一基金）", _headers(stock_columns), _rows(rows["stock_rollup"], stock_columns), widths=_widths(stock_columns))
    industry_columns: list[Column] = [
        ("行业/市场分类", "label", None, 18), ("适用市场", "market", None, 12), ("唯一股票数", "stock_count", "0", 12),
        ("涉及基金数", "fund_count", "0", 12), ("涉及经理数", "manager_count", "0", 12), ("基金-股票记录数", "record_count", "0", 15),
        ("持仓市值合计(万元)", "market_value_sum", "#,##0.00", 19), ("净值比例合计(非加权)", "nav_ratio_sum", "0.00%", 19),
        ("行业快照日期", "snapshot_date", "yyyy-mm-dd", 14), ("说明", "note", None, 35),
    ]
    add_table_sheet(workbook, "申万行业汇总", "公司口径行业与市场暴露汇总（唯一基金）", _headers(industry_columns), _rows(rows["industry_rollup"], industry_columns), widths=_widths(industry_columns))
    anomaly_columns: list[Column] = [
        ("目标基金经理", "manager", None, 14), ("类型", "type", None, 12), ("级别", "severity", None, 10), ("分类", "category", None, 18),
        ("基金/股票代码", "code", "@", 14), ("基金/股票名称", "name", None, 22), ("报告期", "report_date", "yyyy-mm-dd", 13),
        ("问题说明", "message", None, 42), ("是否阻断", "blocking", None, 10), ("建议处理", "action", None, 34), ("来源URL", "source_url", None, 28),
    ]
    add_table_sheet(workbook, "异常与排除", "异常、提示、产品排除与行业时点限制", _headers(anomaly_columns), _rows(rows["anomalies"], anomaly_columns), widths=_widths(anomaly_columns))
    source_rows = [
        ["S01", "天天基金基金经理档案", "完整历史任职基金池及报告期切片", "https://fund.eastmoney.com/manager/{基金经理ID}.html"],
        ["S02", "东方财富基金持仓接口", "季度前十大股票持仓公开披露", "https://fundf10.eastmoney.com/FundArchivesDatas.aspx"],
        ["S03", "申万行业公开快照", "A股申万一级/二级分类；非历史时点", "https://legulegu.com/stockdata/sw-industry-overview"],
        ["R01", "公司去重口径", "跨经理按基金代码＋股票代码去重，共同管理同一基金不重复计算", str(input_path)],
        ["R02", "模型边界", "全部计算由确定性规则完成，未调用DeepSeek", ""],
    ]
    add_table_sheet(workbook, "来源与口径", "数据来源、业务规则与审计说明", ["来源ID", "来源/口径", "用途与限制", "URL或路径"], source_rows, widths={"用途与限制": 55, "URL或路径": 55})
    overview_end = len(rows["overview"]) + 2
    company_end = len(rows["company_holdings"]) + 2
    manager_end = len(rows["manager_holdings"]) + 2
    metrics = data.get("metrics", {})
    add_summary_sheet(
        workbook, f"{label} {quarter} 基金经理前十大持仓分析",
        [("基金经理数", f"=COUNTA('基金经理概览'!$A$3:$A${overview_end})"),
         ("有适用产品经理", f'=COUNTIF(\'基金经理概览\'!$C$3:$C${overview_end},"有适用产品")'),
         ("无适用产品经理", f'=COUNTIF(\'基金经理概览\'!$C$3:$C${overview_end},"无适用产品")'),
         ("报告期在任份额", f"=SUM('基金经理概览'!$D$3:$D${overview_end})"), ("纳入份额", f"=SUM('基金经理概览'!$E$3:$E${overview_end})"),
         ("经理-产品数", f"=SUM('基金经理概览'!$F$3:$F${overview_end})")],
        [("公司唯一基金数", len({row.get("fund_code") for row in rows["company_holdings"]})),
         ("经理口径持仓行", f"=COUNTA('经理持仓明细'!$A$3:$A${manager_end})"),
         ("公司唯一基金持仓行", f"=COUNTA('公司唯一基金持仓'!$A$3:$A${company_end})"),
         ("唯一重仓股票", len(rows["stock_rollup"])), ("唯一A股", metrics.get("global_unique_a_stock_count", 0)),
         ("已匹配申万A股", metrics.get("global_unique_a_stock_mapped", 0)), ("A股行业覆盖率", metrics.get("global_a_industry_coverage", 0))],
        note=f"报告期：{report_date}；数据快照：{data.get('as_of','')}。经理视角按经理－代表产品保留；公司视角按基金代码＋股票代码去重，避免共同管理的同一基金重复计算。行业为当前公开快照，未调用DeepSeek。",
    )
    _move_summary_first(workbook)
    return _save(workbook, output_path)


def build_company_resource_report(input_path: Path, output_path: Path) -> Path:
    data = read_json(input_path)
    summary = data["summary"]
    workbook = _new_workbook()
    company = str(summary.get("company", "基金公司"))
    label = _company_label(company)
    quarter = str(summary.get("quarter", ""))
    overview_columns: list[Column] = [
        ("基金经理", "manager", None, 14), ("天天基金经理ID", "manager_id", "@", 16), ("状态", "status", None, 14),
        ("行业需求数", "industry_demand_count", "0", 12), ("公司需求数", "company_demand_count", "0", 12), ("匹配记录数", "match_count", "0", 12),
        ("原候选匹配数", "source_candidate_match_count", "0", 14), ("业务已确认候选数", "confirmed_candidate_match_count", "0", 16),
        ("待确认候选数", "candidate_match_count", "0", 14), ("待补充数", "pending_count", "0", 12),
        ("排除公司数", "excluded_non_sw_company_count", "0", 12), ("错误说明", "error", None, 32),
    ]
    add_table_sheet(workbook, "基金经理概览", f"{label} {quarter} 基金经理研究资源概览", _headers(overview_columns), _rows(data.get("manager_overview", []), overview_columns), widths=_widths(overview_columns))
    industry_columns: list[Column] = [
        ("最高优先级", "priority", None, 12), ("申万一级行业", "sw_level1", None, 16), ("覆盖经理数", "manager_count", "0", 12),
        ("基金经理", lambda r, _: "、".join(r.get("managers", [])) if isinstance(r.get("managers"), list) else r.get("managers", ""), None, 28),
        ("P1经理数", "p1_manager_count", "0", 11), ("经理需求次数", "demand_occurrences", "0", 12),
        ("不同基金代码数", "unique_fund_count", "0", 14), ("基金代码", lambda r, _: "、".join(r.get("fund_codes", [])) if isinstance(r.get("fund_codes"), list) else r.get("fund_codes", ""), None, 34),
        ("持仓记录算术合计", "holding_count_sum", "0", 16), ("披露市值算术合计(万元)", "market_value_10k_sum", "#,##0.00", 21),
        ("净值比例算术合计", "nav_ratio_sum", "0.00%", 17),
    ]
    add_table_sheet(workbook, "行业需求汇总", f"{label} {quarter} 行业研究资源需求", _headers(industry_columns), _rows(data.get("industry_rollup", []), industry_columns), widths=_widths(industry_columns))
    company_columns: list[Column] = [
        ("最高优先级", "priority", None, 12), ("股票代码", "stock_code", "@", 13), ("股票名称", "stock_name", None, 16),
        ("申万一级行业", "sw_level1", None, 16), ("申万二级行业", "sw_level2", None, 18), ("覆盖经理数", "manager_count", "0", 12),
        ("基金经理", lambda r, _: "、".join(r.get("managers", [])) if isinstance(r.get("managers"), list) else r.get("managers", ""), None, 28),
        ("P1经理数", "p1_manager_count", "0", 11), ("经理需求次数", "demand_occurrences", "0", 12),
        ("不同基金代码数", "unique_fund_count", "0", 14), ("基金代码", lambda r, _: "、".join(r.get("fund_codes", [])) if isinstance(r.get("fund_codes"), list) else r.get("fund_codes", ""), None, 34),
        ("持仓出现次数算术合计", "holding_occurrences_sum", "0", 18), ("披露市值算术合计(万元)", "market_value_10k_sum", "#,##0.00", 21),
        ("单基金最高净值比例", "max_nav_ratio", "0.00%", 18), ("最高匹配分", "best_match_score", "0", 12),
        ("匹配状态", "match_status", None, 16), ("建议对接人员", lambda r, _: "、".join(r.get("matched_people", [])) if isinstance(r.get("matched_people"), list) else r.get("matched_people", ""), None, 28),
    ]
    add_table_sheet(workbook, "公司需求汇总", f"{label} {quarter} 重仓公司研究资源需求", _headers(company_columns), _rows(data.get("company_rollup", []), company_columns), widths=_widths(company_columns))
    person_columns: list[Column] = [
        ("人员姓名", "person_name", None, 14), ("所属机构", "organization", None, 22), ("职务", "job_title", None, 16),
        ("原研究分组", "source_group", None, 18), ("专长标签", "expertise_tags", None, 24), ("地区", "region", None, 12),
        ("覆盖经理数", "manager_count", "0", 12), ("基金经理", lambda r, _: "、".join(r.get("managers", [])) if isinstance(r.get("managers"), list) else r.get("managers", ""), None, 28),
        ("行业目标数", "industry_target_count", "0", 12), ("公司目标数", "company_target_count", "0", 12),
        ("公司精确匹配次数", "exact_company_match_count", "0", 15), ("二级行业匹配次数", "level2_match_count", "0", 15),
        ("业务已确认候选次数", "confirmed_candidate_match_count", "0", 17), ("待确认候选次数", "candidate_match_count", "0", 15),
        ("最高匹配分", "max_score", "0", 12), ("匹配方式", lambda r, _: "、".join(r.get("match_types", [])) if isinstance(r.get("match_types"), list) else r.get("match_types", ""), None, 28),
        ("联系权限", "contact_permission", None, 12), ("联系方式", "contact_info", None, 24),
    ]
    add_table_sheet(workbook, "人员对接汇总", f"{label} {quarter} 研究人员对接汇总", _headers(person_columns), _rows(data.get("person_rollup", []), person_columns), widths=_widths(person_columns))
    match_columns: list[Column] = [
        ("基金经理", "manager", None, 14), ("需求类型", "demand_type", None, 12), ("优先级", "priority", None, 10),
        ("目标代码", "target_code", "@", 14), ("目标名称", "target_name", None, 18), ("申万一级行业", "sw_level1", None, 16),
        ("申万二级行业", "sw_level2", None, 18), ("匹配方式", "match_type", None, 20), ("匹配分", "score", "0", 10),
        ("人员姓名", "person_name", None, 14), ("所属机构", "organization", None, 22), ("职务", "job_title", None, 16),
        ("原研究分组", "source_group", None, 18), ("专长标签", "expertise_tags", None, 24), ("地区", "region", None, 12),
        ("确认状态", "confirmation_status", None, 14), ("原确认状态", "original_confirmation_status", None, 14),
        ("确认人/来源", "confirmed_by", None, 18), ("确认时间（北京时间）", "confirmed_at_beijing", None, 22),
        ("联系权限", "contact_permission", None, 12), ("联系方式", "contact_info", None, 24),
    ]
    add_table_sheet(workbook, "匹配明细", f"{label} {quarter} 全量研究资源匹配明细", _headers(match_columns), _rows(data.get("match_details", []), match_columns), widths=_widths(match_columns))
    candidates = data.get("candidate_items", []) or data.get("confirmed_candidate_items", [])
    add_table_sheet(workbook, "待确认事项", f"{label} {quarter} 原宽口径候选", _headers(match_columns), _rows(candidates, match_columns), widths=_widths(match_columns))
    pending_columns: list[Column] = [
        ("基金经理", "manager", None, 14), ("优先级", "priority", None, 10), ("需求类型", "demand_type", None, 12),
        ("目标代码", "target_code", "@", 14), ("目标名称", "target_name", None, 18), ("申万一级行业", "sw_level1", None, 16),
        ("未匹配原因", "reason", None, 36), ("建议处理", "action", None, 34),
    ]
    add_table_sheet(workbook, "待补充项", f"{label} {quarter} 尚未匹配的研究资源需求", _headers(pending_columns), _rows(data.get("pending_items", []), pending_columns), widths=_widths(pending_columns))
    excluded_columns: list[Column] = [
        ("股票代码", "stock_code", "@", 13), ("股票名称", "stock_name", None, 16), ("市场", "market", None, 10),
        ("行业状态", "sw_level1", None, 14), ("覆盖经理数", "manager_count", "0", 12),
        ("基金经理", lambda r, _: "、".join(r.get("managers", [])) if isinstance(r.get("managers"), list) else r.get("managers", ""), None, 28),
        ("不同基金代码数", "unique_fund_count", "0", 14), ("基金代码", lambda r, _: "、".join(r.get("fund_codes", [])) if isinstance(r.get("fund_codes"), list) else r.get("fund_codes", ""), None, 32),
        ("持仓出现次数算术合计", "holding_occurrences_sum", "0", 18), ("排除原因", "reason", None, 42),
    ]
    add_table_sheet(workbook, "不纳入资源匹配", f"{label} {quarter} 港股及申万不适用公司排除审计", _headers(excluded_columns), _rows(data.get("excluded_rollup", []), excluded_columns), widths=_widths(excluded_columns))
    expected = data.get("checks_expected", {})
    check_specs = [
        ("经理数", "manager_count", "基金经理概览"), ("完成经理数", "completed_manager_count", "运行摘要"),
        ("行业需求数", "industry_demand_count_sum", "行业需求汇总"), ("公司需求数", "company_demand_count_sum", "公司需求汇总"),
        ("匹配记录数", "match_count_sum", "匹配明细"), ("原候选匹配数", "source_candidate_match_count_sum", "待确认事项"),
        ("业务已确认候选数", "confirmed_candidate_match_count_sum", "待确认事项"), ("待确认候选数", "candidate_match_count_sum", "待确认事项"),
        ("待补充数", "pending_count_sum", "待补充项"), ("排除公司数", "excluded_non_sw_company_count_sum", "不纳入资源匹配"),
    ]
    check_rows = []
    for index, (label_name, key, source) in enumerate(check_specs, start=3):
        actual = summary.get(key, 0)
        check_rows.append([label_name, actual, expected.get(key, actual), f'=B{index}-C{index}', f'=IF(D{index}=0,"OK","FAIL")', f"与{source}及回填摘要勾稽"])
    check_ws = add_table_sheet(workbook, "数据校验", f"{label} {quarter} 聚合校验", ["校验项目", "实际值", "预期值", "差异", "状态", "说明"], check_rows)
    check_ws.conditional_formatting.add(f"E3:E{check_ws.max_row}", FormulaRule(formula=['E3="OK"'], fill=PatternFill("solid", fgColor=GREEN)))
    source_rows = [[r.get("item", ""), r.get("rule", ""), "", ""] for r in data.get("rules", [])]
    source_rows.extend([["公司范围", "只汇总同一基金公司、同一报告期的基金经理资源匹配结果", "", ""],
                        ["人员库", str(data.get("personnel_file", "")), str(data.get("personnel_sha256", "")), "联系方式按权限隐藏"],
                        ["模型边界", "全部匹配与汇总由确定性规则完成", "", "未调用DeepSeek"]])
    add_table_sheet(workbook, "来源与口径", f"{label} {quarter} 来源与口径", ["项目", "确定性规则", "审计值", "说明"], source_rows, widths={"确定性规则": 65, "审计值": 45, "说明": 30})
    overview_end = len(data.get("manager_overview", [])) + 2
    add_summary_sheet(
        workbook, f"{label} {quarter} 研究资源对接汇总",
        [("报告期", summary.get("report_date")), ("基金经理数", f"=COUNTA('基金经理概览'!$A$3:$A${overview_end})"),
         ("完成经理数", summary.get("completed_manager_count", 0)), ("行业需求（逐经理合计）", summary.get("industry_demand_count_sum", 0)),
         ("公司需求（逐经理合计）", summary.get("company_demand_count_sum", 0)), ("公司唯一行业数", summary.get("unique_industry_count", 0)),
         ("公司唯一A股数", summary.get("unique_company_count", 0)), ("匹配记录数", summary.get("match_count_sum", 0)),
         ("原候选匹配数", summary.get("source_candidate_match_count_sum", 0)), ("业务已确认候选数", summary.get("confirmed_candidate_match_count_sum", 0)),
         ("尚待确认候选数", summary.get("candidate_match_count_sum", 0)), ("待补充项", summary.get("pending_count_sum", 0)),
         ("数据校验", "通过" if summary.get("status") == "completed" else summary.get("status", ""))],
        [("需要对接的研究人员", summary.get("matched_person_count", 0)), ("公司精确/二级匹配公司", sum(r.get("match_status") == "已匹配" for r in data.get("company_rollup", []))),
         ("组级推定公司", sum(r.get("match_status") == "候选" for r in data.get("company_rollup", []))),
         ("候选已确认公司", summary.get("confirmed_candidate_match_count_sum", 0)), ("尚待确认公司", summary.get("candidate_match_count_sum", 0)),
         ("待补人员公司", summary.get("pending_count_sum", 0)), ("人员联系方式", "全部按权限控制"), ("模型调用", "未调用 DeepSeek")],
        note="公司级报告用于安排研究资源。公司唯一行业/股票按代码去重；需求数、匹配数、市值和净值比例来自逐基金经理结果算术合计，不代表公司统一组合暴露。",
        right_title="决策指标",
    )
    _move_summary_first(workbook)
    return _save(workbook, output_path)


def build_preproduction_replay_report(input_path: Path, output_path: Path) -> Path:
    data = read_json(input_path)
    metrics = data.get("metrics", {})
    workbook = _new_workbook()
    company_columns: list[Column] = [
        ("基金公司", "company", None, 24), ("上期经理数", "previous_manager_count", "0", 12), ("本期经理数", "current_manager_count", "0", 12),
        ("上期正式基金", "previous_formal_funds", "0", 13), ("本期正式基金", "current_formal_funds", "0", 13),
        ("上期公司去重持仓", "previous_holding_rows", "0", 15), ("本期公司去重持仓", "current_holding_rows", "0", 15),
        ("上期移除重复", "previous_duplicate_rows_removed", "0", 13), ("本期移除重复", "current_duplicate_rows_removed", "0", 13),
        ("公司并集", "company_union_count", "0", 11), ("新进", "new_company_count", "0", 9), ("退出", "exited_company_count", "0", 9),
        ("增持", "increased_company_count", "0", 9), ("减持", "decreased_company_count", "0", 9), ("持平", "unchanged_company_count", "0", 9),
        ("行业并集", "industry_union_count", "0", 11), ("去重冲突", "dedup_conflict_count", "0", 11), ("状态", "status", None, 16),
    ]
    add_table_sheet(workbook, "公司比较", "基金公司两期比较", _headers(company_columns), _rows(data.get("company_rows", []), company_columns), widths=_widths(company_columns), title_row=False)
    manager_columns: list[Column] = [
        ("基金公司", "company", None, 24), ("基金经理", "manager", None, 14), ("天天基金经理ID", "manager_id", "@", 16),
        ("Q1状态", "previous_status", None, 14), ("Q2状态", "current_status", None, 14), ("Q1正式基金", "previous_formal_funds", "0", 12),
        ("Q2正式基金", "current_formal_funds", "0", 12), ("基金数变化", "fund_change", "0", 12),
        ("Q1持仓行", "previous_holding_rows", "0", 11), ("Q2持仓行", "current_holding_rows", "0", 11),
        ("新进公司", "new_company_count", "0", 11), ("退出公司", "exited_company_count", "0", 11),
        ("增持公司", "increased_company_count", "0", 11), ("减持公司", "decreased_company_count", "0", 11),
        ("持平公司", "unchanged_company_count", "0", 11), ("比较状态", "comparison_status", None, 18),
        ("管道警告", "pipeline_warning_count", "0", 11), ("行业警告", "industry_warning_count", "0", 11),
    ]
    add_table_sheet(workbook, "经理回放明细", "基金经理两期回放明细", _headers(manager_columns), _rows(data.get("manager_rows", []), manager_columns), widths=_widths(manager_columns), title_row=False)
    add_table_sheet(workbook, "数据质量", "准生产回放数据质量", ["检查ID", "检查项", "实际值", "预期值", "状态", "说明"], [[r.get("check_id"), r.get("check_name"), r.get("actual"), r.get("expected"), r.get("status"), r.get("note")] for r in data.get("checks", [])], widths={"说明": 50}, title_row=False)
    source_rows = [[r.get("item"), "本地审计输入", r.get("path"), r.get("sha256"), ""] for r in data.get("source_files", [])]
    source_rows.extend([["回放限制", "业务限制", limitation, "", ""] for limitation in data.get("limitations", [])])
    add_table_sheet(workbook, "来源与口径", "来源与口径", ["项目", "类型", "本地文件或说明", "SHA-256", "公开来源"], source_rows, widths={"本地文件或说明": 70, "SHA-256": 45}, title_row=False)
    add_summary_sheet(
        workbook, f"{data.get('previous_quarter','')} → {data.get('current_quarter','')} 基金持仓 Agent 全量准生产回放",
        [("回放状态", data.get("overall_status")), ("上期报告日", data.get("previous_report_date")), ("本期报告日", data.get("current_report_date")),
         ("基金经理", metrics.get("manager_count", 0)), ("季度任务", metrics.get("quarter_task_count", 0)),
         ("逐经理比较", metrics.get("manager_comparison_count", 0)), ("通过检查", data.get("passed_check_count", 0)),
         ("失败检查", data.get("failed_check_count", 0)), ("DeepSeek调用", "否")],
        [("Q1逐经理持仓", metrics.get("previous_formal_holding_rows_individual_sum", 0)),
         ("Q1公司去重持仓", metrics.get("previous_company_dedup_holding_rows", 0)),
         ("Q1移除共同管理重复", metrics.get("previous_joint_management_duplicate_rows_removed", 0)),
         ("Q2逐经理持仓", metrics.get("current_formal_holding_rows_individual_sum", 0)),
         ("Q2公司去重持仓", metrics.get("current_company_dedup_holding_rows", 0)),
         ("Q2移除共同管理重复", metrics.get("current_joint_management_duplicate_rows_removed", 0)),
         ("管道错误", metrics.get("pipeline_error_count", 0)), ("行业错误", metrics.get("industry_error_count", 0))],
        note="全量回放只使用本地已披露季度结果，不联网、不调用DeepSeek。公司层已消除共同管理基金重复披露；行业方向可比，但仍保留历史时点限制。",
        right_title="持仓勾稽",
    )
    _move_summary_first(workbook)
    return _save(workbook, output_path)


def build_three_quarter_brief_report(input_path: Path, output_path: Path) -> Path:
    """Build the two-sheet, user-facing three-quarter holdings brief."""
    data = read_json(input_path)
    summary = data.get("summary", {})
    quarters = list(summary.get("quarters", []))
    if len(quarters) != 3:
        raise ValueError("三季度简报输入必须包含连续三个季度")

    workbook = _new_workbook()
    ws = workbook.create_sheet("01_三季持仓")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D6"
    last_col = 3 + len(quarters) * 2
    last_letter = get_column_letter(last_col)

    ws.merge_cells(f"A1:{last_letter}1")
    title = f"{summary.get('company', '')}｜{summary.get('manager', '')}｜最近三季度前十大A股持仓"
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{last_letter}2")
    ws["A2"] = (
        f"报告期：{'、'.join(quarters)}　｜　基础产品 {summary.get('product_count', 0)} 只　｜　"
        "口径：A/C/E去重、只展示A股、保留原披露排名且不递补；空白原因见“99_说明异常”"
    )
    ws["A2"].fill = PatternFill("solid", fgColor=GRAY)
    ws["A2"].font = Font(name="Arial", size=10, color=TEXT)
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    base_headers = ["基金经理", "基础产品名称", "排名"]
    quarter_fills = ["F4B183", "9DC3E6", "A9D18E"]
    for index, header in enumerate(base_headers, start=1):
        ws.merge_cells(start_row=4, start_column=index, end_row=5, end_column=index)
        cell = ws.cell(4, index, header)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=LIGHT_BORDER)
    for index, quarter in enumerate(quarters):
        start_col = 4 + index * 2
        ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=start_col + 1)
        group = ws.cell(4, start_col, quarter)
        group.fill = PatternFill("solid", fgColor=quarter_fills[index])
        group.font = Font(name="Arial", size=11, bold=True, color=TEXT)
        group.alignment = Alignment(horizontal="center", vertical="center")
        for col, label in ((start_col, f"{quarter}持仓"), (start_col + 1, f"{quarter}申万一级行业")):
            cell = ws.cell(5, col, label)
            cell.fill = PatternFill("solid", fgColor=quarter_fills[index])
            cell.font = Font(name="Arial", size=10, bold=True, color=TEXT)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=LIGHT_BORDER)
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 32

    rows = data.get("rows", [])
    for row_index, row in enumerate(rows, start=6):
        values: list[Any] = [row.get("manager", ""), row.get("product_name", ""), row.get("rank", "")]
        quarter_values = row.get("quarters", {})
        for quarter in quarters:
            holding = quarter_values.get(quarter) or {}
            values.extend([holding.get("stock_name", ""), holding.get("sw_level1", "")])
        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row_index, column_index, _excel_value(value))
            cell.font = Font(name="Arial", size=10, color=TEXT)
            cell.alignment = Alignment(
                horizontal="center" if column_index == 3 else "left",
                vertical="center",
                wrap_text=False,
            )
            cell.border = Border(bottom=LIGHT_BORDER)
        if ((row_index - 6) // 10) % 2 == 1:
            for cell in ws[row_index]:
                cell.fill = PatternFill("solid", fgColor="F7FAFC")
        if int(row.get("rank", 0) or 0) == 1:
            for cell in ws[row_index]:
                cell.border = Border(top=Side(style="medium", color="AAB7C4"), bottom=LIGHT_BORDER)
    data_end = max(5, 5 + len(rows))
    ws.auto_filter.ref = f"A5:{last_letter}{data_end}"
    widths = {1: 14, 2: 25, 3: 8}
    for index in range(len(quarters)):
        widths[4 + index * 2] = 17
        widths[5 + index * 2] = 19
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:5"
    ws.auto_filter.ref = f"A5:{last_letter}{data_end}"

    note_ws = workbook.create_sheet("99_说明异常")
    note_ws.sheet_view.showGridLines = False
    note_ws.freeze_panes = "A15"
    _style_title(note_ws, "三季度简报口径、质量与异常", 8)
    notes = [
        ("基金公司", summary.get("company", "")),
        ("基金经理", summary.get("manager", "")),
        ("覆盖季度", "、".join(quarters)),
        ("基础产品数", summary.get("product_count", 0)),
        ("简报A股持仓条数", summary.get("a_share_holding_rows", 0)),
        ("排除非A股持仓条数", summary.get("non_a_holding_rows_excluded", 0)),
        ("空白季度排名格", summary.get("empty_quarter_rank_cells", 0)),
        ("DeepSeek API", "未调用" if not summary.get("deepseek_used") else "已调用"),
    ]
    for row_index, (label, value) in enumerate(notes, start=3):
        note_ws.cell(row_index, 1, label)
        note_ws.cell(row_index, 2, _excel_value(value, label))
        note_ws.cell(row_index, 1).font = Font(name="Arial", bold=True, color=TEXT)
        note_ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=PALE)
        for col in (1, 2):
            note_ws.cell(row_index, col).border = Border(bottom=LIGHT_BORDER)
            note_ws.cell(row_index, col).alignment = Alignment(vertical="center", wrap_text=True)
    note_ws.merge_cells("D3:H10")
    note_ws["D3"] = (
        "阅读规则\n"
        "1. 每个基础产品固定显示排名1—10。\n"
        "2. A/C/E等份额只保留正式代表份额；完整份额仍在单季度底稿中。\n"
        "3. 港股及其他非A股不进入本简报，也不由后续名次递补。\n"
        "4. 空白可能表示产品当季不在管、无相应披露、该排名为非A股或抓取异常。\n"
        "5. 申万行业沿用各季度运行时的行业快照，并非严格历史时点分类。"
    )
    note_ws["D3"].fill = PatternFill("solid", fgColor=GRAY)
    note_ws["D3"].font = Font(name="Arial", size=10, color=TEXT)
    note_ws["D3"].alignment = Alignment(vertical="top", wrap_text=True)

    run_headers = ["季度", "报告期", "基础产品", "正式持仓", "简报A股", "排除非A股", "行业覆盖率", "行业快照日"]
    run_header_row = 13
    for col, header in enumerate(run_headers, start=1):
        note_ws.cell(run_header_row, col, header)
    _style_header(note_ws[run_header_row])
    for row_index, run in enumerate(data.get("quarter_runs", []), start=run_header_row + 1):
        values = [
            run.get("quarter", ""), run.get("report_date", ""), run.get("base_product_count", 0),
            run.get("formal_holding_rows", 0), run.get("a_share_rows_in_brief", 0),
            run.get("non_a_rows_excluded", 0), run.get("industry_coverage", 0), run.get("industry_snapshot_date", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = note_ws.cell(row_index, col, _excel_value(value, run_headers[col - 1]))
            cell.border = Border(bottom=LIGHT_BORDER)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        note_ws.cell(row_index, 7).number_format = "0.00%"
        note_ws.cell(row_index, 2).number_format = "yyyy-mm-dd"
        note_ws.cell(row_index, 8).number_format = "yyyy-mm-dd"

    issue_header_row = run_header_row + max(1, len(data.get("quarter_runs", []))) + 3
    issue_headers = ["季度", "级别", "分类", "代码", "名称", "问题说明", "建议处理", "来源URL"]
    for col, header in enumerate(issue_headers, start=1):
        note_ws.cell(issue_header_row, col, header)
    _style_header(note_ws[issue_header_row])
    issues = data.get("issues", [])
    display_issues = issues or [{"quarter": "全部", "severity": "info", "category": "无异常", "message": "未发现需要人工处理的异常"}]
    for row_index, issue in enumerate(display_issues, start=issue_header_row + 1):
        values = [issue.get(key, "") for key in ("quarter", "severity", "category", "code", "name", "message", "action", "source_url")]
        for col, value in enumerate(values, start=1):
            cell = note_ws.cell(row_index, col, _excel_value(value))
            cell.border = Border(bottom=LIGHT_BORDER)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 4:
                cell.number_format = "@"
            if col == 8 and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
        if str(issue.get("severity", "")).lower() in {"error", "错误"}:
            for cell in note_ws[row_index]:
                cell.fill = PatternFill("solid", fgColor="FCE8E6")

    source_header_row = issue_header_row + len(display_issues) + 3
    source_headers = ["来源ID", "来源/口径", "用途", "URL"]
    for col, header in enumerate(source_headers, start=1):
        note_ws.cell(source_header_row, col, header)
    _style_header(note_ws[source_header_row][:4])
    for row_index, source in enumerate(data.get("sources", []), start=source_header_row + 1):
        values = [source.get("source_id", ""), source.get("name", ""), source.get("usage", ""), source.get("url", "")]
        for col, value in enumerate(values, start=1):
            cell = note_ws.cell(row_index, col, value)
            cell.border = Border(bottom=LIGHT_BORDER)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 4 and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
    for col, width in {1: 15, 2: 24, 3: 22, 4: 16, 5: 20, 6: 44, 7: 36, 8: 48}.items():
        note_ws.column_dimensions[get_column_letter(col)].width = width

    workbook.active = 0
    return _save(workbook, output_path)


def audit_workbook(path: Path, *, expected_sheets: Sequence[str] | None = None) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    formulas = []
    formula_errors = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formulas.append(f"{ws.title}!{cell.coordinate}")
                    if any(error in str(cell.value) for error in FORMULA_ERRORS):
                        formula_errors.append(f"{ws.title}!{cell.coordinate}:{cell.value}")
    missing = [name for name in expected_sheets or [] if name not in workbook.sheetnames]
    return {
        "path": str(path.resolve()), "valid": not missing and not formula_errors,
        "sheet_count": len(workbook.sheetnames), "sheets": workbook.sheetnames,
        "missing_sheets": missing, "formula_count": len(formulas), "formula_errors": formula_errors,
        "size_bytes": path.stat().st_size,
    }
